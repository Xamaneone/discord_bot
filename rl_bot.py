from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font

import discord

from discord.ext import commands

import os

import secret

PREFIX = '.'

PERMISSIONS = (
    "add_reactions", "administrator", "attach_files", "ban_members", "change_nickname", "connect",
    "create_instant_invite",
    "deafen_members", "embed_links", "external_emojis", "kick_members", "manage_channels", "manage_emojis",
    "manage_guild",
    "manage_messages", "manage_nicknames", "manage_permissions", "manage_roles", "manage_webhooks", "mention_everyone",
    "move_members", "mute_members", "priority_speaker", "read_message_history", "read_messages", "request_to_speak",
    "send_messages", "send_tts_messages", "speak", "stream", "use_external_emojis", "use_slash_commands",
    "use_voice_activation", "view_audit_log", "view_channel", "view_guild_insights")

Red = PatternFill(patternType='solid', fgColor='c7182f')

Green = PatternFill(patternType='solid', fgColor='12db3d')

from os import environ

print("Turning on")

FOLDER_PATH = f"C:/Users/{environ.get('USERNAME')}/PycharmProjects/discord_bot"

os.chdir(FOLDER_PATH)

# def is_server_manager(ctx):
#     user_roles = ctx.author.roles
#
#     for role in user_roles:
#         if int(role.id) == 802970241951858738:
#             return True
#     return False


class MyClient(commands.Bot):
    def __init__(self, activity):
        super().__init__(command_prefix=PREFIX, activity=activity)


        @self.command(name="get_data", pass_context=True)
        async def get_data(ctx):

            if not ctx.author.guild_permissions.administrator:
                return

            workbook = Workbook()

            channels = ctx.guild.channels
            # sheet = workbook.active
            for channel in channels:
                sheet = workbook.create_sheet(channel.name)
                # sheet = workbook.active
                # sheet.title = channel.name
                sheet.freeze_panes = "B2"

                roles = channel.overwrites

                for i in range(len(PERMISSIONS)):
                    cell = sheet.cell(row=1, column=i + 2)
                    cell.value = PERMISSIONS[i]
                    cell.font = Font(bold=True, size=10)

                row = 2
                column = 1
                for name, role in roles.items():
                    cell = sheet.cell(row=row, column=column)
                    cell.value = name.name
                    cell.font = Font(bold=True, size=10)
                    for permission in PERMISSIONS:
                        boolean = eval(f"role.{permission}")
                        cell = sheet.cell(row=row, column=column + 1)
                        if boolean is True:
                            boolean = "True"
                            cell.fill = Green
                        else:
                            boolean = "False"
                            cell.fill = Red
                        cell.value = boolean
                        column += 1
                    column -= len(PERMISSIONS)
                    row += 1

                ## Adjusting the columns width
                for col in sheet.columns:
                    max_length = 0
                    column = col[0].column_letter  # Get the column name
                    for cell in col:
                        try:  # Necessary to avoid error on empty cells
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = (max_length + 2) * 1.3
                    sheet.column_dimensions[column].width = adjusted_width


            # roles = ctx.guild.roles
            #
            # for i in range(len(PERMISSIONS)):
            #     cell = sheet.cell(row=1, column=i + 2)
            #     cell.value = PERMISSIONS[i]
            #     cell.font = Font(bold=True, size=10)
            #
            # row = 2
            # column = 1
            # for role in roles:
            #     cell = sheet.cell(row=row, column=column)
            #     cell.value = role.name
            #     cell.font = Font(bold=True, size=10)
            #     for permission in PERMISSIONS:
            #         boolean = eval(f"role.permissions.{permission}")
            #         cell = sheet.cell(row=row, column=column + 1)
            #         if boolean == 1:
            #             boolean = "True"
            #             cell.fill = Green
            #         else:
            #             boolean = "False"
            #             cell.fill = Red
            #         cell.value = boolean
            #         column += 1
            #     column -= len(PERMISSIONS)
            #     row += 1
            #
            # ## Adjusting the columns width
            # for col in sheet.columns:
            #     max_length = 0
            #     column = col[0].column_letter  # Get the column name
            #     for cell in col:
            #         try:  # Necessary to avoid error on empty cells
            #             if len(str(cell.value)) > max_length:
            #                 max_length = len(str(cell.value))
            #         except:
            #             pass
            #     adjusted_width = (max_length + 2) * 1.3
            #     sheet.column_dimensions[column].width = adjusted_width
            #
            #
            #
            sheet = workbook.active
            workbook.remove_sheet(sheet)


            print("DONE")
            workbook.save("server_roles_data.xlsx")
            await ctx.send(file=discord.File("server_roles_data.xlsx"))

    async def on_ready(self):
        print("I'm active!")
        print('Username: {0.name}\nID: {0.id}'.format(self.user))


# activity = discord.Game(name='')
activity = discord.Activity(name='Rizzo', type=discord.ActivityType.watching)

client = MyClient(activity=activity)

client.remove_command('help')

client.run(secret.token_rl)
